from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from race_parser import ParsedRaceFile, Race, Runner


TITLE_FILL = "173B45"
RACE_FILL = "2E6F7E"
HEADER_FILL = "DCECEF"
TEXT = "1F2933"
WHITE = "FFFFFF"


def safe_sheet_title(value: str, used: set[str]) -> str:
    cleaned = "".join(char if char not in r'[]:*?/\\' else " " for char in value).strip()
    cleaned = " ".join(cleaned.split()) or "Output"
    base = cleaned[:31]
    candidate = base
    index = 2
    while candidate in used:
        suffix = f" {index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def sorted_runners(runners: list[Runner]) -> list[Runner]:
    return sorted(
        runners,
        key=lambda runner: (
            runner.figure is None,
            runner.figure if runner.figure is not None else 10_000,
            runner.name,
        ),
    )


def write_race_table(ws, race: Race, start_row: int, start_col: int) -> int:
    runners = sorted_runners(race.runners)
    end_col = start_col + 3

    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row,
        end_column=end_col,
    )
    title_cell = ws.cell(start_row, start_col)
    title_cell.value = race.label
    title_cell.fill = PatternFill("solid", fgColor=RACE_FILL)
    title_cell.font = Font(bold=True, color=WHITE)
    title_cell.alignment = Alignment(vertical="center")

    headers = ["Runner", "Weight lbs", "Last Speed", "Figure"]
    for idx, header in enumerate(headers):
        cell = ws.cell(start_row + 1, start_col + idx)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(bold=True, color=TITLE_FILL)
        cell.alignment = Alignment(vertical="center")

    for row_offset, runner in enumerate(runners, start=2):
        row = start_row + row_offset
        values = [
            runner.name,
            runner.weight_lbs,
            runner.last_speed_figure,
            runner.figure,
        ]
        for col_offset, value in enumerate(values):
            cell = ws.cell(row, start_col + col_offset)
            cell.value = value
            cell.font = Font(color=TEXT)
            cell.alignment = Alignment(vertical="top")

    return len(runners) + 2


def autosize_columns(ws) -> None:
    fixed_widths = {
        "A": 34,
        "B": 12,
        "C": 13,
        "D": 10,
        "E": 3,
        "F": 34,
        "G": 12,
        "H": 13,
        "I": 10,
    }
    for col, width in fixed_widths.items():
        ws.column_dimensions[col].width = width


def write_output_sheet(ws, parsed: ParsedRaceFile, title: str) -> None:
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    ws["A1"] = title
    ws["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    ws["A1"].font = Font(bold=True, color=WHITE, size=14)
    ws["A1"].alignment = Alignment(vertical="center")

    ws.merge_cells("A2:I2")
    ws["A2"] = "Sorted by Weight lbs - Last Speed Figure, lowest first."
    ws["A2"].fill = PatternFill("solid", fgColor="F2F6F7")
    ws["A2"].font = Font(color=TITLE_FILL)

    row = 4
    col = 1
    left_height = 0

    for race in parsed.races:
        height = write_race_table(ws, race, row, col)
        if col == 1:
            left_height = height
            col = 6
        else:
            row += max(left_height, height) + 3
            col = 1
            left_height = 0

    if col == 6:
        row += left_height + 3

    autosize_columns(ws)
    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.4)


def write_audit_sheet(wb: Workbook, parsed_files: list[ParsedRaceFile], used_titles: set[str]) -> None:
    audit = wb.create_sheet(safe_sheet_title("Source Audit", used_titles))
    audit.sheet_view.showGridLines = False
    audit.append(
        [
            "Source PDF",
            "Race",
            "Runner",
            "Age",
            "Weight",
            "Weight lbs",
            "Last Speed",
            "Figure",
            "Last Form Line",
            "Performance Lines Parsed",
        ]
    )
    for cell in audit[1]:
        cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
        cell.font = Font(bold=True, color=WHITE)

    for parsed in parsed_files:
        for race in parsed.races:
            for runner in sorted_runners(race.runners):
                audit.append(
                    [
                        parsed.source_name,
                        race.label,
                        runner.name,
                        runner.age,
                        runner.weight,
                        runner.weight_lbs,
                        runner.last_speed_figure,
                        runner.figure,
                        runner.last_form_line,
                        len(runner.performance_lines),
                    ]
                )

    for idx, width in enumerate([30, 44, 34, 8, 10, 12, 12, 10, 96, 22], start=1):
        audit.column_dimensions[get_column_letter(idx)].width = width
    audit.freeze_panes = "A2"


def build_combined_workbook(parsed_files: list[ParsedRaceFile]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()

    for parsed in parsed_files:
        sheet_title = "Daily Output" if len(parsed_files) == 1 else Path(parsed.source_name).stem
        ws = wb.create_sheet(safe_sheet_title(sheet_title, used_titles))
        display_title = (
            "Race Figure Output"
            if len(parsed_files) == 1
            else f"{Path(parsed.source_name).stem} - Race Figure Output"
        )
        write_output_sheet(ws, parsed, display_title)

    write_audit_sheet(wb, parsed_files, used_titles)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def build_workbook(parsed: ParsedRaceFile) -> bytes:
    return build_combined_workbook([parsed])


def output_filename(source_name: str) -> str:
    stem = Path(source_name).stem or "race-output"
    safe = "".join(char if char.isalnum() or char in (" ", "-", "_") else "-" for char in stem)
    return f"{safe.strip()} - Race Figures.xlsx"


def combined_output_filename(source_names: list[str]) -> str:
    if len(source_names) == 1:
        return output_filename(source_names[0])
    return f"Race Figures - {len(source_names)} PDFs.xlsx"
